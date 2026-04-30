import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# ── RUTAS ──────────────────────────────────────────────────────────────────────
ARCHIVO_CLIENTES = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\clientes (6).xlsx"
ARCHIVO_REPORTE  = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\reporte_camara_consolidado_full_sergio1.xlsx"
ARCHIVO_SALIDA   = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\clientes_actualizado_12.xlsx"

def forzar_texto_nit(valor):
    """Convierte cualquier valor (123.0, '123', 123) en un texto limpio '123'"""
    if valor is None: return ""
    try:
        return str(int(float(str(valor).strip())))
    except:
        return str(valor).strip()

def actualizar_base_clientes():
    print("1. Cargando Reporte Consolidado...")
    df_rep = pd.read_excel(ARCHIVO_REPORTE, dtype=str)
    
    mapa_datos = {}
    for _, fila in df_rep.iterrows():
        nit_limpio = forzar_texto_nit(fila['NIT'])
        if nit_limpio:
            mapa_datos[nit_limpio] = {
                'info': str(fila['INFO CLAVE']) if pd.notna(fila['INFO CLAVE']) else "",
                'camara': str(fila['CAMARA DE COMERCIO']) if pd.notna(fila['CAMARA DE COMERCIO']) else "",
                'firma': str(fila['FIRMA']) if pd.notna(fila['FIRMA']) else ""
            }
    
    print(f"   -> {len(mapa_datos)} NITs listos para transferir.")

    print("2. Abriendo Excel de Clientes...")
    wb_cli = load_workbook(ARCHIVO_CLIENTES)
    ws_cli = wb_cli.active

    # Localizar columnas en Clientes
    indices = {}
    for cell in ws_cli[1]:
        if cell.value:
            header = str(cell.value).strip().upper()
            if header == "NIT": indices['nit'] = cell.column
            if "CLAVE CAMARA Y COMERCIO" in header: indices['col_info'] = cell.column
            if "CAMARA DE COMERCIO" in header: indices['col_camara'] = cell.column
            if "CLAVE PARA FIRMAR CAMARA Y COMERCIO" in header: indices['col_firma'] = cell.column

    # Verificación de columna NIT
    if 'nit' not in indices:
        print("❌ Error: No encontré la columna 'NIT' en Clientes.")
        return

    # Crear columnas si no existen
    if 'col_info' not in indices:
        indices['col_info'] = ws_cli.max_column + 1
        ws_cli.cell(1, indices['col_info'], "CLAVE CAMARA Y COMERCIO")
    if 'col_camara' not in indices:
        indices['col_camara'] = ws_cli.max_column + 1
        ws_cli.cell(1, indices['col_camara'], "Camara de Comercio")
    if 'col_firma' not in indices:
        indices['col_firma'] = ws_cli.max_column + 1
        ws_cli.cell(1, indices['col_firma'], "CLAVE PARA FIRMAR CAMARA Y COMERCIO")

    print("3. Ejecutando cruce de datos...")
    contador = 0
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for r in range(2, ws_cli.max_row + 1):
        raw_nit_cli = ws_cli.cell(r, indices['nit']).value
        nit_cliente = forzar_texto_nit(raw_nit_cli)

        if nit_cliente in mapa_datos:
            datos = mapa_datos[nit_cliente]
            
            # 1. Actualizar Clave Cámara y Comercio
            c1 = ws_cli.cell(r, indices['col_info'], datos['info'])
            
            # 2. Actualizar Cámara de Comercio
            c2 = ws_cli.cell(r, indices['col_camara'], datos['camara'])
            
            # 3. Actualizar CLAVE PARA FIRMAR CAMARA Y COMERCIO (Nueva)
            c3 = ws_cli.cell(r, indices['col_firma'], datos['firma'])
            
            # Aplicar estilos a las 3 celdas
            for celda in [c1, c2, c3]:
                celda.alignment = Alignment(wrap_text=True, vertical="top")
                celda.border = borde
            
            contador += 1

    if contador > 0:
        wb_cli.save(ARCHIVO_SALIDA)
        print(f"✅ ¡Éxito total! Se actualizaron {contador} clientes.")
        print(f"Archivo guardado: {ARCHIVO_SALIDA}")
    else:
        print("❌ El cruce falló. No se encontraron coincidencias de NIT.")

if __name__ == "__main__":
    actualizar_base_clientes()