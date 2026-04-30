import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

# ── RUTAS ──────────────────────────────────────────────────────────────────────
ARCHIVO_CLIENTES = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\clientes (6).xlsx"
ARCHIVO_CAMARA   = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\CLAVES CAMARA DE COMERCIO SERGIO.xlsx"
SALIDA           = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\reporte_camara_consolidado_full_sergio1.xlsx"

# ── ESTILOS ────────────────────────────────────────────────────────────────────
FILL_HEADER = PatternFill("solid", start_color="2F5496", end_color="2F5496")
FONT_HEADER = Font(bold=True, color="FFFFFF")
BORDE = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

def extraer_valor_firma(texto):
    """Busca la palabra 'firma' y extrae el valor que le sigue"""
    if not texto or texto == "Sin comentarios":
        return ""
    lineas = texto.split('\n')
    for linea in lineas:
        if 'firma' in linea.lower():
            # Busca lo que hay después de 'firma:'
            match = re.search(r'firma\s*:\s*(.*)', linea, re.IGNORECASE)
            if match:
                return match.group(1).strip()
    return ""

def ejecutar_proceso():
    print("1. Cargando base de Clientes...")
    try:
        df_cli = pd.read_excel(ARCHIVO_CLIENTES, dtype=str)
        col_nit_cli = next((c for c in df_cli.columns if 'nit' in str(c).lower()), None)
        nits_clientes_set = set(df_cli[col_nit_cli].str.strip().tolist()) if col_nit_cli else set()
    except:
        nits_clientes_set = set()

    print("2. Procesando Libro de Cámara...")
    wb_cam = load_workbook(ARCHIVO_CAMARA)
    ws_cam = wb_cam.active

    indices = {}
    for cell in ws_cam[1]: 
        if cell.value:
            nombre = str(cell.value).strip().lower()
            if 'responsable' in nombre: indices['responsable'] = cell.column
            if nombre == 'a' or nombre == 'columnaa': indices['columna_a'] = cell.column
            if 'nit' in nombre: indices['nit'] = cell.column
            if 'cliente' in nombre: indices['cliente'] = cell.column
            if 'renovado' in nombre: indices['renovado'] = cell.column
            if 'pagado' in nombre: indices['pagado'] = cell.column
            if 'certificados' in nombre: indices['certificados'] = cell.column
            if 'camara' in nombre and 'comercio' in nombre: indices['camara'] = cell.column
            if 'clave' in nombre and 'renovacion' in nombre: indices['clave_col'] = cell.column

    COLUMNA_NOTAS = 9 
    datos_acumulados = {}

    for r in range(2, ws_cam.max_row + 1):
        nit_val = str(ws_cam.cell(r, indices['nit']).value).strip() if ws_cam.cell(r, indices['nit']).value else ""
        if not nit_val or nit_val == "None": continue

        # Clave física
        valor_clave_celda = ""
        if 'clave_col' in indices:
            val = ws_cam.cell(r, indices['clave_col']).value
            valor_clave_celda = str(val).strip() if val else ""

        # Comentario de la columna I
        celda_con_nota = ws_cam.cell(r, COLUMNA_NOTAS)
        texto_comentario = celda_con_nota.comment.text.strip() if celda_con_nota.comment else "Sin comentarios"
        
        # Extraer firma de este comentario individual
        firma_individual = extraer_valor_firma(texto_comentario)

        info_fila = {
            'responsable': ws_cam.cell(r, indices['responsable']).value,
            'columna_a': ws_cam.cell(r, indices.get('columna_a', 1)).value,
            'nit': nit_val,
            'cliente': ws_cam.cell(r, indices['cliente']).value,
            'renovado': ws_cam.cell(r, indices.get('renovado', 1)).value,
            'pagado': ws_cam.cell(r, indices.get('pagado', 1)).value,
            'certificados': ws_cam.cell(r, indices.get('certificados', 1)).value,
            'entidad': str(ws_cam.cell(r, indices['camara']).value or "N/A"),
            'clave': valor_clave_celda,
            'firma': firma_individual,
            'comentario_original': texto_comentario,
            'estado_en_clientes': "ENCONTRADO" if nit_val in nits_clientes_set else "NO ENCONTRADO"
        }

        if nit_val not in datos_acumulados:
            datos_acumulados[nit_val] = []
        datos_acumulados[nit_val].append(info_fila)

    print("3. Consolidando resultados y asignando firmas por cámara...")
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Consolidado"

    headers = [
        "RESPONSABLE", "ColumnA", "NIT", "CLIENTES", "RENOVADO", 
        "PAGADO", "CERTIFICADOS", "CAMARA DE COMERCIO", "CLAVE RENOVACION", 
        "FIRMA", "ESTADO EN CLIENTES", "INFO CLAVE"
    ]
    
    for col, h in enumerate(headers, 1):
        cell = ws_out.cell(1, col, h)
        cell.font, cell.fill, cell.border = FONT_HEADER, FILL_HEADER, BORDE

    fila_out = 2
    for nit, lista in datos_acumulados.items():
        base = lista[0]
        
        # Consolidación
        camaras_unidas = " | ".join([f["entidad"] for f in lista])
        claves_unidas = " | ".join([f"{f['entidad']}: {f['clave']}" for f in lista])
        
        # FIRMA: Ahora indica a qué cámara pertenece cada una
        # Formato: MEDELLIN: xxxx | ABURRA SUR: yyyy
        firmas_con_entidad = []
        for f in lista:
            if f['firma']:
                firmas_con_entidad.append(f"{f['entidad']}: {f['firma']}")
        
        firmas_finales = " | ".join(firmas_con_entidad)
        
        # INFO CLAVE
        info_unida = "\n".join([f"--- {f['entidad']} ---\n{f['comentario_original']}\n" for f in lista])
        
        # Escritura
        ws_out.cell(fila_out, 1, base['responsable'])
        ws_out.cell(fila_out, 2, base['columna_a'])
        ws_out.cell(fila_out, 3, nit)
        ws_out.cell(fila_out, 4, base['cliente'])
        ws_out.cell(fila_out, 5, base['renovado'])
        ws_out.cell(fila_out, 6, base['pagado'])
        ws_out.cell(fila_out, 7, base['certificados'])
        ws_out.cell(fila_out, 8, camaras_unidas)
        ws_out.cell(fila_out, 9, claves_unidas)
        ws_out.cell(fila_out, 10, firmas_finales)
        ws_out.cell(fila_out, 11, base['estado_en_clientes'])
        ws_out.cell(fila_out, 12, info_unida)

        for c in range(1, 13):
            ws_out.cell(fila_out, c).border = BORDE
            ws_out.cell(fila_out, c).alignment = Alignment(vertical="top", wrap_text=True)
        
        fila_out += 1

    anchos = [20, 10, 15, 30, 12, 12, 12, 30, 30, 35, 20, 60]
    for i, w in enumerate(anchos, 1):
        ws_out.column_dimensions[get_column_letter(i)].width = w

    wb_out.save(SALIDA)
    print(f"✅ Proceso completado exitosamente en: {SALIDA}")

if __name__ == "__main__":
    ejecutar_proceso()