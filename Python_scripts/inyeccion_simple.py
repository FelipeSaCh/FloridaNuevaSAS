import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side

# ── 1. CONFIGURACIÓN DE RUTAS ──────────────────────────────────────────────────
ARCHIVO_CLIENTES = r"D:\Escritorio\MANEJO GLIDE\recursos\clientes6.xlsx"
ARCHIVO_REPORTE  = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\reporte_verificacion_nits_seguridadsocial.xlsx"
ARCHIVO_SALIDA   = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\clientes_actualizado_13.xlsx"

# ── 2. CONFIGURACIÓN DE COLUMNAS (MODIFICA SOLO AQUÍ) ──────────────────────────
# Formato: "Nombre en el archivo reporte": "Nombre deseado en archivo clientes"
COLUMNAS_A_INYECTAR = {
    "USUARIO ISS": "USUARIO ISS",
    "CLAVE ISS": "CLAVE ISS"
}

# Nombre de la columna común para el cruce (en ambos archivos)
COLUMNA_NIT_REPORTE = "NIT"
COLUMNA_NIT_CLIENTES = "NIT"

# ── FUNCIONES DE APOYO ────────────────────────────────────────────────────────
def forzar_texto_nit(valor):
    if valor is None or pd.isna(valor): return ""
    try:
        return str(int(float(str(valor).strip())))
    except:
        return str(valor).strip()

def actualizar_base_clientes():
    print("1. Cargando Reporte y preparando mapa de datos...")
    df_rep = pd.read_excel(ARCHIVO_REPORTE, dtype=str)
    
    mapa_datos = {}
    for _, fila in df_rep.iterrows():
        nit_limpio = forzar_texto_nit(fila[COLUMNA_NIT_REPORTE])
        if nit_limpio:
            # Guardamos un diccionario con los valores de las columnas configuradas
            mapa_datos[nit_limpio] = {original: str(fila[original]) if pd.notna(fila[original]) else "" 
                                     for original in COLUMNAS_A_INYECTAR.keys()}
    
    print(f"   -> {len(mapa_datos)} NITs cargados del reporte.")

    print("2. Abriendo Excel de Clientes...")
    wb_cli = load_workbook(ARCHIVO_CLIENTES)
    ws_cli = wb_cli.active

    # --- Localización de columnas existentes ---
    headers_actuales = {str(cell.value).strip().upper(): cell.column for cell in ws_cli[1] if cell.value}
    
    indices_finales = {}
    
    # Localizar NIT
    if COLUMNA_NIT_CLIENTES.upper() not in headers_actuales:
        print(f"❌ Error: No encontré la columna '{COLUMNA_NIT_CLIENTES}' en Clientes.")
        return
    col_nit_idx = headers_actuales[COLUMNA_NIT_CLIENTES.upper()]

    # Localizar o Crear columnas de destino
    for col_reporte, col_cliente in COLUMNAS_A_INYECTAR.items():
        nombre_busqueda = col_cliente.strip().upper()
        if nombre_busqueda in headers_actuales:
            indices_finales[col_reporte] = headers_actuales[nombre_busqueda]
        else:
            # Si no existe, se crea al final
            nueva_col = ws_cli.max_column + 1
            ws_cli.cell(1, nueva_col, col_cliente)
            indices_finales[col_reporte] = nueva_col
            headers_actuales[nombre_busqueda] = nueva_col # Actualizar mapa interno

    print("3. Ejecutando inyección de datos dinámicos...")
    contador = 0
    borde = Border(left=Side(style="thin"), right=Side(style="thin"), 
                   top=Side(style="thin"), bottom=Side(style="thin"))

    for r in range(2, ws_cli.max_row + 1):
        nit_cliente = forzar_texto_nit(ws_cli.cell(r, col_nit_idx).value)

        if nit_cliente in mapa_datos:
            datos_nuevos = mapa_datos[nit_cliente]
            
            for col_rep, col_cli_idx in indices_finales.items():
                valor_inyectar = datos_nuevos[col_rep]
                celda = ws_cli.cell(r, col_cli_idx, valor_inyectar)
                
                # Estilos automáticos
                celda.alignment = Alignment(wrap_text=True, vertical="top")
                celda.border = borde
            
            contador += 1

    if contador > 0:
        wb_cli.save(ARCHIVO_SALIDA)
        print(f"✅ ¡Proceso completado! Se actualizaron {contador} filas.")
        print(f"Destino: {ARCHIVO_SALIDA}")
    else:
        print("❌ No se encontraron coincidencias para actualizar.")

if __name__ == "__main__":
    actualizar_base_clientes()