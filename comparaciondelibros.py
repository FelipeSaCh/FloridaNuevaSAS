import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuración de Rutas ─────────────────────────────────────────────────────
ARCHIVO_CLIENTES = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\clientes (3).xlsx"
ARCHIVO_CAMARA   = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\CLAVES CAMARA DE COMERCIO SEBASTIAN.xlsx"
SALIDA           = r"D:\Escritorio\MANEJO GLIDE\Comparacionlibros\reporte_camara_consolidado.xlsx"

# Estilos
FILL_HEADER = PatternFill("solid", start_color="2F5496", end_color="2F5496")
FONT_HEADER = Font(bold=True, color="FFFFFF")
BORDE = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

def extraer_clave_nota(celda):
    """Extrae la clave de renovación si existe en el comentario de la celda"""
    if celda and celda.comment:
        lineas = celda.comment.text.split('\n')
        for linea in lineas:
            if 'clave renovacion' in linea.lower():
                return linea.split(':', 1)[1].strip() if ':' in linea else linea
    return ""

def ejecutar_proceso():
    print("1. Leyendo lista de NITs en Clientes para validación...")
    try:
        df_cli = pd.read_excel(ARCHIVO_CLIENTES, dtype=str)
        col_nit_cli = next((c for c in df_cli.columns if 'nit' in str(c).lower()), None)
        nits_clientes_set = set(df_cli[col_nit_cli].str.strip().tolist()) if col_nit_cli else set()
    except Exception as e:
        print(f"⚠️ No se pudo leer Clientes correctamente: {e}")
        nits_clientes_set = set()

    print("2. Procesando libro de Cámara...")
    wb_cam = load_workbook(ARCHIVO_CAMARA, data_only=True)
    ws_cam = wb_cam.active

    indices = {}
    for cell in ws_cam[1]: 
        if cell.value:
            nombre = str(cell.value).strip().lower()
            if 'responsable' in nombre: indices['responsable'] = cell.column
            if 'nit' in nombre: indices['nit'] = cell.column
            if 'cliente' in nombre: indices['cliente'] = cell.column
            if 'camara' in nombre: indices['camara'] = cell.column
            if 'clave' in nombre and 'renovacion' in nombre: indices['clave_col'] = cell.column

    for col_req in ['responsable', 'nit', 'cliente', 'camara']:
        if col_req not in indices:
            print(f"❌ Error: No se encontró la columna '{col_req.upper()}' en Cámara.")
            return

    # Estructura para almacenar y consolidar: { NIT: [lista_de_diccionarios_con_info] }
    datos_acumulados = {}

    print("3. Recolectando y analizando datos...")
    for r in range(2, ws_cam.max_row + 1):
        nit_val = str(ws_cam.cell(r, indices['nit']).value).strip() if ws_cam.cell(r, indices['nit']).value else ""
        if not nit_val or nit_val == "None": continue

        # Extraer Clave
        clave_final = ""
        if 'clave_col' in indices:
            val_col = ws_cam.cell(r, indices['clave_col']).value
            if val_col: clave_final = str(val_col).strip()
        
        if not clave_final:
            clave_final = extraer_clave_nota(ws_cam.cell(r, indices['camara']))

        info_fila = {
            'responsable': ws_cam.cell(r, indices['responsable']).value,
            'cliente': ws_cam.cell(r, indices['cliente']).value,
            'entidad': str(ws_cam.cell(r, indices['camara']).value or "DESCONOCIDA"),
            'clave': clave_final or "SIN CLAVE",
            'estado': "ENCONTRADO" if nit_val in nits_clientes_set else "NO ENCONTRADO"
        }

        if nit_val not in datos_acumulados:
            datos_acumulados[nit_val] = []
        datos_acumulados[nit_val].append(info_fila)

    # 4. Crear archivo de salida y consolidar
    print("4. Generando archivo consolidado...")
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Reporte Consolidado"

    headers = ["RESPONSABLE", "NIT", "CLIENTE", "CAMARA DE COMERCIO", "CLAVE RENOVACION", "ESTADO EN CLIENTES"]
    for col, h in enumerate(headers, 1):
        cell = ws_out.cell(1, col, h)
        cell.font, cell.fill, cell.border = FONT_HEADER, FILL_HEADER, BORDE
        ws_out.column_dimensions[get_column_letter(col)].width = 30

    fila_out = 2
    for nit, lista_filas in datos_acumulados.items():
        if len(lista_filas) > 1:
            # --- CASO REPETIDO: CONSOLIDAR ---
            # Usamos los datos del primer registro para las columnas generales
            base = lista_filas[0]
            
            # Construimos la cadena: CAMARA1: CLAVE1, CAMARA2: CLAVE2
            claves_consolidadas = " | ".join([f"{f['entidad']}: {f['clave']}" for f in lista_filas])
            
            ws_out.cell(fila_out, 1, base['responsable'])
            ws_out.cell(fila_out, 2, nit)
            ws_out.cell(fila_out, 3, base['cliente'])
            ws_out.cell(fila_out, 4, "") # Vacío en Cámara de Comercio según tu instrucción
            ws_out.cell(fila_out, 5, claves_consolidadas)
            ws_out.cell(fila_out, 6, base['estado'])
        else:
            # --- CASO ÚNICO: NORMAL ---
            f = lista_filas[0]
            ws_out.cell(fila_out, 1, f['responsable'])
            ws_out.cell(fila_out, 2, nit)
            ws_out.cell(fila_out, 3, f['cliente'])
            ws_out.cell(fila_out, 4, f['entidad'])
            ws_out.cell(fila_out, 5, f['clave'])
            ws_out.cell(fila_out, 6, f['estado'])

        # Estilo de bordes y alineación
        for c in range(1, 7):
            ws_out.cell(fila_out, c).border = BORDE
            ws_out.cell(fila_out, c).alignment = Alignment(vertical="center", wrap_text=True)
        
        fila_out += 1

    try:
        wb_out.save(SALIDA)
        print(f"✅ Proceso terminado. Archivo consolidado en: {SALIDA}")
    except PermissionError:
        print("❌ Error: El archivo de salida está abierto.")

if __name__ == "__main__":
    ejecutar_proceso()